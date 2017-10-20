import xlwings as xw
import openpyxl

wb1 = xw.Book('sample.xlsx')
wb2 = xw.Book('FileName.xlsx')


for i, o in range(2, wb1.max_column):
    x = wb1.cell(row=o, column = i)
    y = wb2.cell(row=o, column = i)
    if x == y:
        print('Good')
    else:
        print('bad')
