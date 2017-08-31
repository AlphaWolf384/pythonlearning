''' Excel to Excel Data Checker/Updater v0.1 '''

path1 = "C:\Users\Public\Documents\Rockwell\Dom\CVBFY18NonConfigMasterv12-FINAL.xlsx"
path2 = "C:\Users\Public\Documents\Rockwell\Dom\PVPNegotiationToolFinal8-7-17.xlsx"

from openpyxl import load_workbook
import time
# The source xlsx file is named as source.xlsx
wb1=load_workbook("C:\Users\Public\Documents\Rockwell\Dom\CVBFY18NonConfigMasterv12-FINAL.xlsx")
wb2=load_workbook("C:\Users\Public\Documents\Rockwell\Dom\PVPNegotiationToolFinal8-7-17.xlsx")

ws1 = wb1.active
ws2 = wb2.active
first_column1 = ws1['A']
first_column2 = ws2['E']

# Print the contents
while True:
    user = raw_input("Ready? or Press X to stop ")
    if user != "x" or user != "X":
        for x in xrange(len(first_column1)):
            for y in xrange(len(first_column2)):
                col1 = first_column1[x].value
                col2 = first_column2[y].value
                data1 = ws1['C']
                data2 = ws2['H']
                row1 = data1[x].value
                row2 = data2[y].value
                if col1 == col2:
                    if row1 == row2:
                        print(col1 + " & " + col2 + " No needed to Update")
                    else:
                        row2 = row1
                        print(col1 + " & " + col2 + " are updated")

    else:
        time.sleep(3)
        break


    
